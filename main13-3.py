from openpyxl import load_workbook
from openpyxl import Workbook
import requests
import re

url = 'https://n.news.naver.com/article/016/0001999997?cds=news_media_pc&type=editn'

headers = {
    'User-Agent':'Mozilla/5.0',
    'Content-Type':'text/htlm;charset=utf-8'
}

reponse = requests.get(url,headers=headers)
results = re.findall(r'[\w\.-]+@[\w\.-]+',reponse.text)
results = list(set(results))
print(results)

try:
    wb = load_workbook(r'C:\\python40\\email.xlsx',data_only=True)
    sheet = wb.active
except:
    wb = Workbook()
    sheet = wb.active
    
for result in results:
    sheet.append([result])
    
wb.save(r'C:\\python40\\email.xlsx')      