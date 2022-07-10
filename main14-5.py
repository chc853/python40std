from openpyxl import load_workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

load_wb = load_workbook(r".\email.xlsx", data_only=True)
load_ws = load_wb.active

for i in range(1,load_ws.max_row+1):
    recv_email_value = load_ws.cell(i,1).value
    print("email : ",recv_email_value)
    try:
        send_email = "chc853@naver.com"
        send_pwd = "!#zaq1xsw2"
        recv_email = recv_email_value
        
        smtp_name = "smtp.naver.com"
        smtp_port = 587
        
        msg = MIMEMultipart()
        
        msg['Subject'] = "엑셀에서 읽어서 발송하는 메일"
        msg['From'] = send_email
        msg['To'] = recv_email
        text = """
            메일 내용입니다.
            감사합니다.
        """
        msg.attach(MIMEText(text))
        s=smtplib.SMTP(smtp_name,smtp_port)
        s.starttls()
        s.login(send_email,send_pwd)
        s.sendmail(send_email,recv_email,msg.as_string())
        s.quit()
    except:
        print("에러 : ",recv_email_value)
